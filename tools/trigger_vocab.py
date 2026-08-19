"""Build the trigger vocabulary from the sources, so it is data and not code.

    python tools/trigger_vocab.py            # rebuild unittransfer/data/trigger_vocab.json
    python tools/trigger_vocab.py --check    # report what would change, write nothing
    python tools/trigger_vocab.py --audit    # what the reference tool's list gets wrong

The ``Trigger / WhenToTest / Condition / Affects`` language has a few hundred
condition terms and a couple of hundred events, and which of them exist is a
property of the *engine*, not of our code. Baking that list into Python would
mean editing Python every time somebody finds another one — so it is generated
here, from three sources, in this order of authority:

1. **The Docudemons spreadsheet** (``Reference/TWCenter/M2TW_Ultimate_Docudemons_5.3.xlsx``)
   — the community's reference list: 415 conditions and 222 events, each with a
   description, a worked example, and (the valuable part) which data types the
   event *exports* and which the condition *requires*.
2. **Every EDCT and EDA on this machine** — what mods actually write. This is
   where the argument *shape* of each term comes from: the spreadsheet describes
   parameters in prose ("logic token, test value"), whereas 25 000 real condition
   lines say exactly which tokens appear in which order.
3. **The reference tool's ``conditionDefs.jsx``** — consulted, not trusted. It is
   the reason ``--audit`` exists: a good third of its event list does not exist
   in the engine (it invents an ``On`` prefix — ``OnCharacterTurnStart`` for the
   real ``CharacterTurnStart``), and porting it as-is would have put names in our
   dropdowns that no mod can use.

The output ``trigger_vocab.json`` ships in releases; this script does not, and
neither does openpyxl.
"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

XLSX = ROOT / "Reference" / "TWCenter" / "M2TW_Ultimate_Docudemons_5.3.xlsx"
OUT = ROOT / "unittransfer" / "data" / "trigger_vocab.json"
UPSTREAM_DEFS = "refs/upstream/editor/main:src/components/shared/conditionDefs.jsx"

#: the files whose triggers we measure argument shapes from
TRIGGER_FILES = ("export_descr_character_traits.txt", "export_descr_ancillaries.txt")

OPS = ("<=", ">=", "!=", "==", "<", ">", "=")


# ---------------------------------------------------------------------------
# 1. the spreadsheet


def docudemon_records(sheet) -> list:
    """Rows of ``label: value`` pairs, split into records at each ``Identifier:``."""
    out, cur = [], None
    for row in sheet.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        label = (cells[1] if len(cells) > 1 else "").rstrip(":")
        value = " ".join(x for x in cells[2:] if x).strip()
        if label == "Identifier":
            cur = {"Identifier": value}
            out.append(cur)
        elif cur is not None and label:
            cur[label] = value
    return [r for r in out if r.get("Identifier")]


def read_docudemons() -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    conds, events = {}, {}
    for r in docudemon_records(wb["Conditions"]):
        conds[r["Identifier"]] = {
            "term": r["Identifier"],
            "requires": _types(r.get("Trigger requirements", "")),
            "params": _clean(r.get("Parameters", "")),
            "sample": _clean(r.get("Sample use", "")),
            "hint": _clean(r.get("Description", "")),
            "where": _clean(r.get("Battle or Strat", "")),
        }
    for r in docudemon_records(wb["Events"]):
        events[r["Identifier"]] = {
            "event": r["Identifier"],
            "hint": _clean(r.get("Event", "")),
            # an event exports a flat set; only requirements are ever disjunctive
            "exports": _flat(_types(r.get("Exports", ""))),
        }
    attrs = []
    for row in wb["Character Attributes"].iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row and row[0] else ""
        if name and name.lower() != "name":
            attrs.append(name)
    wb.close()
    return {"conditions": conds, "events": events, "attributes": sorted(set(attrs))}


def _clean(s: str) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    return "" if s.lower() in ("none", "n/a") else s


#: The spreadsheet is a community document and spells a few data types more than
#: one way. Left alone, each spelling would read as a type the event does not
#: export and :func:`triggers.check` would cry wolf on a few dozen sound triggers.
ALIASES = {
    "character": "character_record",            # BattleGeneralRouted's Exports
    "nc_character_records": "nc_character_record",   # Religion's requirements
    "target faction": "target_faction",
}


def _norm(t: str) -> str:
    t = re.sub(r"\s+", " ", t.strip().lower())
    return ALIASES.get(t, t)


def _types(s: str) -> list:
    """Requirements/exports as alternative groups: ``[[a, b], [c]]`` = (a∧b) ∨ c.

    Commas mean "all of these"; ``or`` means "any of these". Only one entry in
    the whole spreadsheet uses ``or`` (``Religion``), but reading it as a single
    literal type is the difference between two real findings and a hundred
    false ones.
    """
    s = (s or "").strip()
    if not s or s.lower() in ("none", "n/a"):
        return []
    groups = []
    for alt in re.split(r"\bor\b", s):
        parts = sorted({_norm(t) for t in re.split(r"[,\n]", alt) if t.strip()})
        if parts:
            groups.append(parts)
    return groups


def _flat(groups: list) -> list:
    """Every type named anywhere in the groups — what an Exports list becomes."""
    return sorted({t for g in groups for t in g})


# ---------------------------------------------------------------------------
# 2. what the mods on this machine actually write


def shape_of(tokens: list) -> str:
    """The argument shape of one condition's tokens, as a pattern string.

    ``["Gandalf", ">", "0"]`` -> ``"name op num"``. This is the only part of the
    vocabulary that has to be exact — it decides what boxes the GUI draws — so it
    is measured rather than read off prose.
    """
    out = []
    for t in tokens:
        if t in OPS:
            out.append("op")
        elif _is_num(t):
            out.append("num")
        else:
            out.append("name")
    return " ".join(out)


def _is_num(t: str) -> bool:
    try:
        float(t)
        return True
    except ValueError:
        return False


def scan_mods(mods_root: Path) -> dict:
    """Measured shapes, event counts and vocabulary seen in real trigger files."""
    from unittransfer import triggers

    shapes = defaultdict(Counter)
    events = Counter()
    attributes = Counter()
    files = 0
    for mod in sorted(p for p in mods_root.iterdir() if p.is_dir()):
        for name in TRIGGER_FILES:
            path = mod / "data" / name
            if not path.exists():
                continue
            files += 1
            text = path.read_text(encoding=triggers.ENCODING)
            parsed = triggers.parse_text(text)
            for trig in parsed.triggers:
                if trig.when_to_test:
                    events[trig.when_to_test] += 1
                for cond in trig.conditions:
                    shapes[cond.term][shape_of(cond.args)] += 1
            # `Effect <attribute> <points>` lines, in the trait and ancillary
            # definitions above the triggers. A line scan and not a parse on
            # purpose: this runs over both file shapes and only wants the word.
            for line in text.split("\n"):
                w = line.split(";", 1)[0].split()
                if len(w) == 3 and w[0] == "Effect":
                    attributes[w[1]] += 1
    return {"files": files, "shapes": shapes, "events": events,
            "attributes": attributes}


def reconcile_attributes(documented: list, measured: Counter) -> list:
    """The spreadsheet's attribute list, spelled the way the engine reads it.

    The sheet is a community document and two of its entries are not what a mod
    writes: ``Hitpoints`` for the engine's ``HitPoints`` (515 real ``Effect``
    lines), and ``Chivalry/Dread`` for ``Chivalry``, which is one attribute whose
    negative half the sheet names alongside it. Both are the commonest attributes
    there are, so leaving them alone means the traits editor reports several
    hundred sound ``Effect`` lines as unknown and the builder's picker offers two
    names no mod can use.

    Measurement corrects a spelling; it never adds a name. Mods write one
    ``Combat_V_Faction_<faction>`` per faction, and a list built from those would
    be this machine's mods rather than the engine.
    """
    written = {}
    for name, _ in measured.most_common():
        written.setdefault(name.lower(), name)
    out = []
    for name in documented:
        parts = name.split("/")
        out.extend([written[p.lower()] for p in parts if p.lower() in written]
                   or [parts[0]])
    return sorted(set(out))


# ---------------------------------------------------------------------------
# 3. the reference tool's list, consulted and audited


def upstream_defs() -> dict:
    """``{conditions: [...], events: [...]}`` from ``conditionDefs.jsx``, or empty."""
    try:
        text = subprocess.run(["git", "show", UPSTREAM_DEFS], cwd=ROOT,
                              capture_output=True, text=True, check=True).stdout
    except (OSError, subprocess.CalledProcessError):
        return {"conditions": [], "events": []}
    conds = re.findall(r"\{\s*key:\s*'([^']+)'", text)
    block = re.search(r"WHEN_TO_TEST_OPTIONS\s*=\s*\[(.*?)\]", text, re.S)
    events = re.findall(r"'([^']+)'", block.group(1)) if block else []
    return {"conditions": conds, "events": events}


# ---------------------------------------------------------------------------
# putting it together


def build(mods_root) -> dict:
    doc = read_docudemons()
    seen = scan_mods(mods_root) if mods_root and mods_root.is_dir() else {
        "files": 0, "shapes": {}, "events": Counter(), "attributes": Counter()}

    conditions = []
    for term in sorted(set(doc["conditions"]) | set(seen["shapes"])):
        d = doc["conditions"].get(term, {})
        measured = seen["shapes"].get(term, Counter())
        # the shape the mods on this machine actually write, commonest first;
        # falling back to the spreadsheet's worked example when nothing uses it
        shapes = [s for s, _ in measured.most_common()]
        if not shapes:
            shapes = [_shape_from_sample(term, d.get("sample", ""))]
        conditions.append({
            "term": term,
            "shapes": shapes,
            "uses": sum(measured.values()),
            "requires": d.get("requires", []),
            "hint": d.get("hint", ""),
            "params": d.get("params", ""),
            "sample": d.get("sample", ""),
            "where": d.get("where", ""),
            # a term real mods use that the reference list has never heard of is
            # not an error — it is the reason the parser warns instead of dropping
            "documented": term in doc["conditions"],
        })

    events = []
    for name in sorted(set(doc["events"]) | set(seen["events"])):
        d = doc["events"].get(name, {})
        events.append({
            "event": name,
            "uses": seen["events"].get(name, 0),
            "exports": d.get("exports", []),
            "hint": d.get("hint", ""),
            "documented": name in doc["events"],
        })

    return {
        "version": 1,
        "source": {
            "docudemons": XLSX.name,
            "trigger_files_scanned": seen["files"],
            "conditions_documented": len(doc["conditions"]),
            "events_documented": len(doc["events"]),
        },
        "operators": list(OPS),
        "attributes": reconcile_attributes(doc["attributes"], seen["attributes"]),
        "conditions": conditions,
        "events": events,
    }


def _shape_from_sample(term: str, sample: str) -> str:
    """Best-effort shape for a term no installed mod uses, from its example."""
    toks = (sample or "").split()
    if toks and toks[0] == term:
        toks = toks[1:]
    return shape_of(toks)


def audit(vocab: dict) -> str:
    up = upstream_defs()
    known_c = {c["term"] for c in vocab["conditions"]}
    known_e = {e["event"] for e in vocab["events"]}
    real_c = [t for t in up["conditions"] if t in known_c]
    real_e = [t for t in up["events"] if t in known_e]
    lines = [
        "# Reference-tool trigger vocabulary — audit",
        "",
        "_Generated by `python tools/trigger_vocab.py --audit`. Regenerate after an "
        "upstream sync._",
        "",
        "## Verdict: do not port their list",
        "",
        f"`src/components/shared/conditionDefs.jsx` lists {len(up['conditions'])} "
        f"conditions, of which **{len(real_c)} exist**; and {len(up['events'])} "
        f"`WhenToTest` events, of which **{len(real_e)} exist**. The engine has "
        f"{len(known_c)} conditions and {len(known_e)} events.",
        "",
        "Two failure modes run through it:",
        "",
        "* **Invented `On` prefixes.** The engine's events are `CharacterTurnStart`, "
        "`PostBattle`, `CaptureSettlement`. Theirs are `OnCharacterTurnStart`, "
        "`OnCaptureSettlement` — a naming convention borrowed from somewhere else. "
        "A dropdown built from that list writes triggers that never fire.",
        "* **Plausible-sounding conditions that do not exist** — `IsSpy`, `IsHeir`, "
        "`OwnsWife`, `SettlementLevel`, `GuildLevel`. They read like conditions a "
        "modder would want, which is exactly what makes them dangerous in a picker.",
        "",
        "So our vocabulary is generated from the Docudemons spreadsheet plus every "
        "trigger file on the machine (`tools/trigger_vocab.py`), and their file is "
        "kept only as the reason this audit exists. Nothing in it is ported.",
        "",
        "The lists below are the full comparison.",
        "",
    ]
    for label, theirs, known in (("conditions", up["conditions"], known_c),
                                 ("events", up["events"], known_e)):
        bogus = [t for t in theirs if t not in known]
        lines += [f"## {label}: {len(theirs) - len(bogus)} real, {len(bogus)} unknown to "
                  f"the engine and to every installed mod", ""]
        lines += [f"- `{t}`" for t in bogus] or ["- (none)"]
        lines.append("")
    ours_only = sorted(known_c - set(up["conditions"]))
    lines += [f"## conditions they do not list: {len(ours_only)}", "",
              ", ".join(f"`{t}`" for t in ours_only[:40])
              + (" …" if len(ours_only) > 40 else ""), ""]
    return "\n".join(lines)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--check", action="store_true",
                    help="report what would change, write nothing")
    ap.add_argument("--audit", action="store_true",
                    help="print the reference-tool comparison instead of building")
    ap.add_argument("--mods", default="", help="mods folder to measure shapes from")
    args = ap.parse_args(argv)

    from unittransfer import config
    mods_root = Path(args.mods) if args.mods else None
    if mods_root is None:
        root = config.get_med2_root()
        mods_root = Path(root) / "mods" if root else None

    if not XLSX.exists():
        print(f"missing {XLSX} — the Docudemons spreadsheet is the vocabulary source")
        return 2
    vocab = build(mods_root)
    if args.audit:
        print(audit(vocab))
        return 0

    new = json.dumps(vocab, indent=1, ensure_ascii=False, sort_keys=True) + "\n"
    old = OUT.read_text(encoding="utf-8") if OUT.exists() else ""
    src = vocab["source"]
    print(f"{len(vocab['conditions'])} conditions ({src['conditions_documented']} documented), "
          f"{len(vocab['events'])} events, measured over "
          f"{src['trigger_files_scanned']} trigger file(s)")
    if new == old:
        print(f"{OUT.relative_to(ROOT)} is already up to date")
        return 0
    if args.check:
        print(f"{OUT.relative_to(ROOT)} WOULD CHANGE")
        return 1
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(new, encoding="utf-8")
    print(f"wrote {OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
